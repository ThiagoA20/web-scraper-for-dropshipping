from xlsxwriter import Workbook as wbw
import openpyxl
import time
from shutil import move
from os import getcwd
from convert_currency import to_aed, to_mxn, to_cad, to_inr

counter_rows = 2


def convert(price):
    if price != 0:
        try:
            price = str(price)
            price = [p for p in price]
            price[-1] = 9
            price[-2] = 9
            price_converted = ''
            for i in price:
                price_converted += str(i)
            return float(price_converted)
        except IndexError:
            return float(0)
    else:
        return 0


def spreadsheet():
    title = time.ctime(time.time()).replace(' ', '_').replace(':', '-') + ".xlsx"
    outworkbook = wbw(f"{getcwd()}\{title}")
    outsheet = outworkbook.add_worksheet()

    salmao = outworkbook.add_format(
        {
            "bg_color": "#FCD5B4"
        }
    )

    yellow = outworkbook.add_format(
        {
            "bg_color": "#FFFF00"
        }
    )

    red = outworkbook.add_format(
        {
            "bg_color": "#FF8080"
        }
    )

    columns = {
        "A1": ["Page Url"],
        "B1": ["Product Url"],
        "C1": ["Sku-product"],
        "D1": ["Brand"],
        "E1": ["Name", salmao],
        "F1": ["Description", salmao],
        "G1": ["Bullet Point", salmao],
        "H1": ["Bullet Point", salmao],
        "I1": ["Bullet Point", salmao],
        "J1": ["Bullet Point", salmao],
        "K1": ["Bullet Point", salmao],
        "L1": ["Main image", salmao],
        "M1": ["Variant image", yellow],
        "N1": ["Variant image", yellow],
        "O1": ["Variant image", yellow],
        "P1": ["Variant image", yellow],
        "Q1": ["Parentage", red],
        "R1": ["Parent-sku", red],
        "S1": ["SP INR", salmao],
        "T1": ["MRP INR", salmao],
        "V1": ["Variant Names"],
        "W1": ["SP CAD"],
        "X1": ["MRP CAD"],
        "Y1": ["SP MXN"],
        "Z1": ["MRP MXN"],
        "AA1": ["SP AED"],
        "AB1": ["MRP AED"],
    }
    for key in columns:
        try:
            outsheet.write(key, str(columns[key][0]), columns[key][1])
        except IndexError:
            outsheet.write(key, str(columns[key][0]))

    outworkbook.close()
    return title


def append_spreadsheet(product, driver, title):
    try:
        global counter_rows
        wb_obj = openpyxl.load_workbook(title)
        st_obj = wb_obj.active

        price_inr = to_inr(product, driver)
        price_cad = to_cad(product, driver)
        price_mxn = to_mxn(product, driver)
        price_aed = to_aed(product, driver)

        if price_inr == "Unexpected Error x01" or price_cad == "Unexpected Error x01" or price_mxn == "Unexpected Error x01" or price_aed == "Unexpected Error x01":
            raise Exception("Unexpected Error x01")
        else:
            pass

        st_obj.cell(row=counter_rows, column=1, value=product.page_url)
        st_obj.cell(row=counter_rows, column=2, value=product.url)
        st_obj.cell(row=counter_rows, column=4, value=product.brand)
        st_obj.cell(row=counter_rows, column=3, value=product.sku)
        st_obj.cell(row=counter_rows, column=5, value=product.name)
        st_obj.cell(row=counter_rows, column=6, value=product.description)
        st_obj.cell(row=counter_rows, column=12, value=product.main_image)

        st_obj.cell(row=counter_rows, column=17, value="Parent")

        for i in range(4):
            try:
                st_obj.cell(row=counter_rows, column=13 + i, value=product.other_main_images[i])
            except IndexError:
                pass

        if len(product.price) == 1:
            for i in range(5):
                st_obj.cell(row=counter_rows, column=i + 7, value=product.bullet_points[i])
            for i in range(len(product.price)):
                st_obj.cell(row=counter_rows + i, column=2, value=product.url)
                st_obj.cell(row=counter_rows + i, column=4, value=product.brand)
                if product.store != 3:
                    st_obj.cell(row=counter_rows + i, column=3, value=f"{product.sku}")
                else:
                    st_obj.cell(row=counter_rows + i, column=3, value=f"{product.asin[i]}")
                try:
                    st_obj.cell(row=counter_rows + i + 1, column=12, value=product.variant_images[i])
                except IndexError:
                    st_obj.cell(row=counter_rows + i + 1, column=12, value=" ")
                st_obj.cell(row=counter_rows + i, column=18, value=product.sku)
                st_obj.cell(row=counter_rows + i, column=19, value=convert("%.0f" % price_inr[i]))
                st_obj.cell(row=counter_rows + i, column=20, value=convert("%.0f" % (price_inr[i] * 1.2)))
                try:
                    st_obj.cell(row=counter_rows + i + 1, column=22, value=product.variant_names[i])
                except IndexError:
                    st_obj.cell(row=counter_rows + i + 1, column=22, value=" ")
                st_obj.cell(row=counter_rows + i, column=23, value=convert("%.2f" % price_cad[i]))
                st_obj.cell(row=counter_rows + i, column=24, value=convert("%.2f" % (price_cad[i] * 1.2)))
                st_obj.cell(row=counter_rows + i, column=25, value=convert("%.2f" % price_mxn[i]))
                st_obj.cell(row=counter_rows + i, column=26, value=convert("%.2f" % (price_mxn[i] * 1.2)))
                st_obj.cell(row=counter_rows + i, column=27, value=convert("%.2f" % price_aed[i]))
                st_obj.cell(row=counter_rows + i, column=28, value=convert("%.2f" % (price_aed[i] * 1.2)))
            counter_rows += len(product.price)
        else:
            for i in range(5):
                for i2 in range(len(product.price)):
                    st_obj.cell(row=counter_rows + i2 + 1, column=i + 7, value=product.bullet_points[i])
            for i in range(len(price_inr)):
                st_obj.cell(row=counter_rows + i + 1, column=2, value=product.url)
                st_obj.cell(row=counter_rows + i + 1, column=4, value=product.brand)
                if product.store != 3:
                    st_obj.cell(row=counter_rows + i + 1, column=3, value=f"{product.sku}-{i + 1}")
                else:
                    st_obj.cell(row=counter_rows + i + 1, column=3, value=f"{product.asin[i]}")
                st_obj.cell(row=counter_rows + i + 1, column=5, value=product.name)
                st_obj.cell(row=counter_rows + i + 1, column=6, value=product.description)
                st_obj.cell(row=counter_rows + i + 1, column=17, value="Child")
                try:
                    st_obj.cell(row=counter_rows + i + 1, column=12, value=product.variant_images[i])
                except IndexError:
                    st_obj.cell(row=counter_rows + i + 1, column=12, value=" ")
                st_obj.cell(row=counter_rows + i + 1, column=18, value=product.sku)
                st_obj.cell(row=counter_rows + i + 1, column=19, value=convert("%.0f" % price_inr[i]))
                st_obj.cell(row=counter_rows + i + 1, column=20, value=convert("%.0f" % (price_inr[i] * 1.2)))
                try:
                    st_obj.cell(row=counter_rows + i + 1, column=22, value=product.variant_names[i])
                except IndexError:
                    st_obj.cell(row=counter_rows + i + 1, column=22, value=" ")
                st_obj.cell(row=counter_rows + i + 1, column=23, value=convert("%.2f" % price_cad[i]))
                st_obj.cell(row=counter_rows + i + 1, column=24, value=convert("%.2f" % (price_cad[i] * 1.2)))
                st_obj.cell(row=counter_rows + i + 1, column=25, value=convert("%.2f" % price_mxn[i]))
                st_obj.cell(row=counter_rows + i + 1, column=26, value=convert("%.2f" % (price_mxn[i] * 1.2)))
                st_obj.cell(row=counter_rows + i + 1, column=27, value=convert("%.2f" % price_aed[i]))
                st_obj.cell(row=counter_rows + i + 1, column=28, value=convert("%.2f" % (price_aed[i] * 1.2)))
            counter_rows += len(product.price) + 1
        wb_obj.save(f"{getcwd()}\{title}")
    except:
        raise Exception("Unexpected Error x02")


def close(folder_selected, title):
    global counter_rows
    move(f"{getcwd()}\{title}", folder_selected)
    counter_rows = 2
